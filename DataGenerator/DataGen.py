import openpyxl

def Data_generator():
    wb = openpyxl.load_workbook("E:\Selenium\EndToEndAutomation\DataGenerator\ExcelDataGen.xlsx")
    wk = wb['Sheet1']
    r = wk.max_row
    li = []
    li1 = []
    for i in range(1,r+1):
        sh1 = wk.cell(i,1)
        sh2 = wk.cell(i,2)
        li1=[]
        li1.insert(0,sh1.value)
        li1.insert(1,sh2.value)
        li.insert(i-1,li1)

    return li



